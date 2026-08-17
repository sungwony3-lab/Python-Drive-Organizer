import base64
import json
import os
import sqlite3
import tempfile
import unittest
from pathlib import Path

from docx import Document
from openpyxl import load_workbook

from database import initialize_schema
from search_service import SearchService
from test_search_service import database_hash, seed_database
from tree_export_service import (
    MAX_OPENAI_FILE_BYTES,
    SnapshotInfo,
    TreeCursorError,
    TreeCursorStaleError,
    TreeExportFileTooLargeError,
    TreeExportNotFoundError,
    build_openai_file_response,
    create_tree_export,
    get_snapshot_info,
    paginate_tree,
    resolve_export_file,
)


class TreeExportServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        self.database_path = self.root / "drive_index.db"
        self.export_directory = self.root / "exports"
        self.log_path = self.root / "tree_export.log"
        self.connection = sqlite3.connect(self.database_path)
        self.connection.row_factory = sqlite3.Row
        initialize_schema(self.connection)
        seed_database(self.connection)
        self.connection.execute(
            """
            INSERT INTO scan_state (
                scan_id, status, started_at, finished_at, files_seen,
                folders_seen, scope_type, message, created_at
            ) VALUES (
                'SCAN-TREE', 'COMPLETED', '2026-08-17T00:00:00Z',
                '2026-08-17T00:01:00Z', 6, 10, 'USER_DRIVE', NULL,
                '2026-08-17T00:00:00Z'
            )
            """
        )
        self.connection.commit()
        self.secret = "S" * 48

    def tearDown(self) -> None:
        self.connection.close()
        self.temporary_directory.cleanup()

    def full_tree(self, **kwargs):
        return SearchService(self.connection).build_full_tree(**kwargs)

    def export(self, export_format: str, **kwargs):
        return create_tree_export(
            self.connection,
            export_format=export_format,
            root_folder=kwargs.get("root_folder"),
            include_files=kwargs.get("include_files", True),
            max_depth=kwargs.get("max_depth"),
            export_directory=self.export_directory,
            audit_log_path=self.log_path,
        )

    def test_full_tree_counts_order_and_determinism(self) -> None:
        first = self.full_tree(include_files=True)
        second = self.full_tree(include_files=True)
        folder_only = self.full_tree(include_files=False)

        self.assertEqual(first.items, second.items)
        self.assertEqual(len(first.items), 16)
        self.assertEqual(len(first.folder_ids), 10)
        self.assertEqual(len(first.file_ids), 6)
        self.assertEqual(len(folder_only.items), 10)
        self.assertEqual(folder_only.file_ids, set())
        ids = [item["id"] for item in first.items]
        self.assertEqual(len(ids), len(set(ids)))
        a_children = [
            item["id"]
            for item in first.items
            if item["parent_id"] == "A"
        ]
        self.assertEqual(a_children, ["B", "D", "S1", "f5"])

    def test_subtree_and_max_depth(self) -> None:
        subtree = self.full_tree(
            root_folder_id="B", include_files=True, max_depth=1
        )
        self.assertEqual(
            [item["id"] for item in subtree.items],
            ["B", "C", "f6"],
        )
        self.assertTrue(all(item["level"] <= 1 for item in subtree.items))

    def test_pagination_has_no_missing_or_duplicate_nodes(self) -> None:
        tree = self.full_tree(include_files=True)
        snapshot = get_snapshot_info(self.connection)
        cursor = None
        collected = []
        while True:
            page = paginate_tree(
                tree,
                snapshot,
                root_folder=None,
                include_files=True,
                max_depth=None,
                page_size=3,
                cursor=cursor,
                secret=self.secret,
            )
            collected.extend(item["id"] for item in page["items"])
            cursor = page["next_cursor"]
            if cursor is None:
                self.assertFalse(page["has_more"])
                break
        expected = [item["id"] for item in tree.items]
        self.assertEqual(collected, expected)
        self.assertEqual(len(collected), len(set(collected)))

    def test_cursor_rejects_tampering_options_and_stale_snapshot(self) -> None:
        tree = self.full_tree(include_files=True)
        snapshot = get_snapshot_info(self.connection)
        first = paginate_tree(
            tree,
            snapshot,
            root_folder=None,
            include_files=True,
            max_depth=None,
            page_size=2,
            cursor=None,
            secret=self.secret,
        )
        cursor = first["next_cursor"]
        with self.assertRaises(TreeCursorError):
            paginate_tree(
                tree,
                snapshot,
                root_folder=None,
                include_files=True,
                max_depth=None,
                page_size=2,
                cursor=cursor[:-1] + ("A" if cursor[-1] != "A" else "B"),
                secret=self.secret,
            )
        with self.assertRaises(TreeCursorError):
            paginate_tree(
                tree,
                snapshot,
                root_folder="A",
                include_files=True,
                max_depth=None,
                page_size=2,
                cursor=cursor,
                secret=self.secret,
            )
        changed_snapshot = SnapshotInfo(
            latest_scan_id="SCAN-CHANGED",
            latest_scan_status="COMPLETED",
            scan_finished_at=snapshot.scan_finished_at,
            indexed_files=snapshot.indexed_files,
            indexed_folders=snapshot.indexed_folders,
        )
        with self.assertRaises(TreeCursorStaleError):
            paginate_tree(
                tree,
                changed_snapshot,
                root_folder=None,
                include_files=True,
                max_depth=None,
                page_size=2,
                cursor=cursor,
                secret=self.secret,
            )

    def test_txt_docx_and_xlsx_exports_open_and_match_counts(self) -> None:
        before = database_hash(self.connection)
        txt = self.export("txt")
        docx = self.export("docx")
        xlsx = self.export("xlsx")
        after = database_hash(self.connection)

        self.assertEqual(before, after)
        for result in (txt, docx, xlsx):
            self.assertEqual(result.node_count, 16)
            self.assertEqual(result.folder_count, 10)
            self.assertEqual(result.file_count, 6)
            self.assertEqual(result.latest_scan_id, "SCAN-TREE")
            path, _media_type = resolve_export_file(
                result.export_id, self.export_directory
            )
            self.assertGreater(path.stat().st_size, 0)

        txt_path, _ = resolve_export_file(txt.export_id, self.export_directory)
        txt_content = txt_path.read_text(encoding="utf-8")
        self.assertIn("한글 보고서.pdf", txt_content)
        self.assertIn("SQLite snapshot", txt_content)

        docx_path, _ = resolve_export_file(docx.export_id, self.export_directory)
        document = Document(docx_path)
        document_text = "\n".join(p.text for p in document.paragraphs)
        self.assertIn("Google Drive Tree", document_text)
        self.assertIn("한글 보고서.pdf", document_text)

        xlsx_path, _ = resolve_export_file(xlsx.export_id, self.export_directory)
        workbook = load_workbook(xlsx_path, read_only=False, data_only=True)
        try:
            sheet = workbook["Drive Tree"]
            self.assertEqual(sheet["A10"].value, "Level")
            self.assertEqual(sheet.freeze_panes, "A11")
            self.assertEqual(sheet.auto_filter.ref, "A10:H26")
            self.assertEqual(sheet.max_row, 26)
        finally:
            workbook.close()

    def test_audit_contains_no_tree_names_and_export_id_is_opaque(self) -> None:
        result = self.export("txt")
        log_text = self.log_path.read_text(encoding="utf-8")
        self.assertEqual(len(result.export_id), 32)
        self.assertIn(result.export_id, log_text)
        self.assertNotIn("한글 보고서", log_text)
        self.assertNotIn("Alpha Plan", log_text)
        event = json.loads(log_text.splitlines()[-1])
        self.assertEqual(
            set(event),
            {"timestamp", "export_id", "format", "node_count", "status"},
        )

    def test_invalid_or_unknown_export_id_is_not_resolved(self) -> None:
        with self.assertRaises(TreeExportNotFoundError):
            resolve_export_file("../secret", self.export_directory)
        with self.assertRaises(TreeExportNotFoundError):
            resolve_export_file("A" * 32, self.export_directory)

    def test_openai_file_response_matches_all_export_formats(self) -> None:
        expected_mime_types = {
            "txt": "text/plain",
            "docx": (
                "application/vnd.openxmlformats-officedocument."
                "wordprocessingml.document"
            ),
            "xlsx": (
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        }
        for export_format, expected_mime_type in expected_mime_types.items():
            with self.subTest(export_format=export_format):
                result = self.export(export_format)
                path, _ = resolve_export_file(
                    result.export_id, self.export_directory
                )
                response = build_openai_file_response(
                    result.export_id,
                    export_directory=self.export_directory,
                    audit_log_path=self.log_path,
                )
                self.assertEqual(set(response), {"openaiFileResponse"})
                self.assertEqual(len(response["openaiFileResponse"]), 1)
                item = response["openaiFileResponse"][0]
                self.assertEqual(item["name"], path.name)
                self.assertEqual(item["mime_type"], expected_mime_type)
                self.assertEqual(base64.b64decode(item["content"]), path.read_bytes())

        log_text = self.log_path.read_text(encoding="utf-8")
        self.assertNotIn("openaiFileResponse", log_text)
        self.assertNotIn("content", log_text)

    def test_openai_file_response_rejects_unknown_traversal_and_over_10mb(
        self,
    ) -> None:
        for export_id in ("../secret", "A" * 32):
            with self.subTest(export_id=export_id):
                with self.assertRaises(TreeExportNotFoundError):
                    build_openai_file_response(
                        export_id,
                        export_directory=self.export_directory,
                        audit_log_path=self.log_path,
                    )

        self.export_directory.mkdir(parents=True, exist_ok=True)
        export_id = "Z" * 32
        oversized_path = (
            self.export_directory / f"drive_tree_test_{export_id}.txt"
        )
        with oversized_path.open("wb") as handle:
            handle.truncate(MAX_OPENAI_FILE_BYTES + 1)
        with self.assertRaises(TreeExportFileTooLargeError):
            build_openai_file_response(
                export_id,
                export_directory=self.export_directory,
                audit_log_path=self.log_path,
            )
        last_event = json.loads(
            self.log_path.read_text(encoding="utf-8").splitlines()[-1]
        )
        self.assertEqual(last_event["status"], "GPT_FILE_TOO_LARGE")
        self.assertEqual(last_event["byte_size"], MAX_OPENAI_FILE_BYTES + 1)
        self.assertNotIn("content", last_event)

    def test_default_paths_are_independent_of_working_directory(self) -> None:
        from tree_export_service import EXPORT_AUDIT_LOG_PATH, EXPORT_DIRECTORY

        original = Path.cwd()
        try:
            os.chdir(self.root)
            self.assertTrue(EXPORT_DIRECTORY.is_absolute())
            self.assertTrue(EXPORT_AUDIT_LOG_PATH.is_absolute())
        finally:
            os.chdir(original)


if __name__ == "__main__":
    unittest.main()
